import os
import pandas as pd
import re
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

# ================= 🔧 [ কনফিগারেশন ] =================
# রেলওয়ে ভেরিয়েবল থেকে টোকেন নেওয়া
BOT_TOKEN = os.environ.get("BOT_TOKEN", "8260254278:AAE0ZTPrPVQExDHS0VWhA7T8f_Bp8S1gYiI")

def auto_adjust_column_width(file_path):
    """Excel ফাইলের কলামের প্রস্থ অটো সাইজ করে"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    except:
        pass

def extract_from_single_column(df):
    """যদি মাত্র ১টি কলাম থাকে, তাহলে স্পেস দিয়ে আলাদা করা ডাটা এক্সট্রাক্ট করে"""
    all_data = []
    
    # শুধু প্রথম কলাম নেওয়া
    for idx, row in df.iterrows():
        if pd.notna(row[0]):
            text = str(row[0]).strip()
            if text:
                # স্পেস দিয়ে আলাদা করা
                parts = text.split()
                
                # যতগুলো অংশ আছে সব নেওয়া (সর্বোচ্চ ৫টি)
                row_data = parts[:5]
                
                # ৫টি কলাম পূরণ করা
                while len(row_data) < 5:
                    row_data.append("")
                
                all_data.append(row_data)
    
    if all_data:
        columns = ["col1", "col2", "col3", "col4", "col5"]
        df_clean = pd.DataFrame(all_data, columns=columns)
        
        # খালি ডাটা বাদ
        df_clean = df_clean[(df_clean["col1"].astype(str).str.strip() != "") & 
                            (df_clean["col2"].astype(str).str.strip() != "")]
        
        # col2 (password) এবং col1 (username) অনুযায়ী সাজানো
        df_clean = df_clean.sort_values(by=["col2", "col1"])
        
        return df_clean
    
    return None

def extract_from_columns(df):
    """একাধিক কলাম থেকে ডাটা নেয়া (সর্বোচ্চ ৫টি)"""
    all_data = []
    
    # কতগুলো কলাম আছে দেখি
    num_cols = df.shape[1]
    
    # যদি ১টি কলাম থাকে, তাহলে single column ফাংশন কল করি
    if num_cols == 1:
        return extract_from_single_column(df)
    
    # সর্বোচ্চ ৫টি কলাম নেওয়া
    max_cols = min(5, num_cols)
    
    for idx, row in df.iterrows():
        row_data = []
        
        # প্রতিটি কলাম থেকে ডাটা নেওয়া
        for col_idx in range(max_cols):
            if pd.notna(row[col_idx]):
                # যদি কোনো কলামে স্পেস দিয়ে একাধিক ডাটা থাকে
                text = str(row[col_idx]).strip()
                if text and ' ' in text:
                    # স্পেস দিয়ে আলাদা করে সবগুলো অংশ নেওয়া
                    parts = text.split()
                    row_data.extend(parts[:5])  # প্রথম ৫টি অংশ নেওয়া
                else:
                    row_data.append(text)
            else:
                row_data.append("")
        
        # যদি row_data তে ৫টির বেশি ডাটা থাকে, তাহলে প্রথম ৫টি নেওয়া
        if len(row_data) > 5:
            row_data = row_data[:5]
        
        # ৫টি কলাম পূরণ করা
        while len(row_data) < 5:
            row_data.append("")
        
        # যদি col1 এবং col2 থাকে
        if row_data[0] and row_data[1]:
            all_data.append(row_data)
    
    if all_data:
        columns = ["col1", "col2", "col3", "col4", "col5"]
        df_clean = pd.DataFrame(all_data, columns=columns)
        
        # খালি ডাটা বাদ
        df_clean = df_clean[(df_clean["col1"].astype(str).str.strip() != "") & 
                            (df_clean["col2"].astype(str).str.strip() != "")]
        
        # col2 (password) এবং col1 (username) অনুযায়ী সাজানো
        df_clean = df_clean.sort_values(by=["col2", "col1"])
        
        return df_clean
    
    return None

def extract_from_messy_data(df):
    """এলোমেলো ডাটা থেকে ৫টি কলাম খুঁজে বের করা"""
    all_data = []
    
    # পুরো ফাইল স্ক্যান করে ডাটা খোঁজা
    for idx, row in df.iterrows():
        for col_idx, val in enumerate(row):
            if pd.notna(val):
                text = str(val).strip()
                if text:
                    # চেষ্টা 1: স্পেস দিয়ে আলাদা কিনা
                    parts = text.split()
                    
                    if len(parts) >= 2:
                        # ২টি বা তার বেশি কলাম থাকলে
                        row_data = parts[:5]  # প্রথম ৫টি নেওয়া
                        while len(row_data) < 5:
                            row_data.append("")
                        all_data.append(row_data)
                        break  # একটি সেল থেকে ডাটা পেলে ব্রেক করি
    
    if all_data:
        columns = ["col1", "col2", "col3", "col4", "col5"]
        df_clean = pd.DataFrame(all_data, columns=columns)
        
        # খালি ডাটা বাদ
        df_clean = df_clean[(df_clean["col1"].astype(str).str.strip() != "") & 
                            (df_clean["col2"].astype(str).str.strip() != "")]
        
        # col2 এবং col1 অনুযায়ী সাজানো
        df_clean = df_clean.sort_values(by=["col2", "col1"])
        
        return df_clean
    
    return None

# ================= 🚀 [ স্টার্ট কমান্ড ] =================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "✅ *Duplicate Remover Bot Active!*\n\n"
        "📂 Send me a `.csv` or `.xlsx` file.\n\n"
        "🔧 *What this bot does:*\n"
        "• Works with 1 or more columns\n"
        "• If 1 column: splits data by spaces\n"
        "• Extracts up to 5 columns (A, B, C, D, E)\n"
        "• Ignores extra columns beyond E\n"
        "• Removes duplicate entries (checks all columns)\n"
        "• Groups same passwords together\n\n"
        "📤 Output: `col1 | col2 | col3 | col4 | col5` (sorted by col2/password)\n\n"
        "💡 Powered by MAX FUTURE",
        parse_mode="Markdown"
    )

# ================= 📂 [ ফাইল প্রসেসিং ] =================
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file_name = update.message.document.file_name
    input_file = f"input_{file_name}"
    unique_file = "Unique_Data.xlsx"
    duplicate_file = "Duplicate_Data.xlsx"

    try:
        # ফাইল ডাউনলোড
        file = await context.bot.get_file(update.message.document.file_id)
        await file.download_to_drive(input_file)

        await update.message.reply_text("⏳ Processing file...")

        # ফাইল লোড
        if file_name.lower().endswith('.csv'):
            try:
                df = pd.read_csv(input_file, encoding='utf-8', header=None)
            except:
                df = pd.read_csv(input_file, encoding='latin1', header=None)
        else:
            df = pd.read_excel(input_file, header=None)

        # প্রথমে কলাম থেকে এক্সট্রাক্ট করার চেষ্টা
        clean_df = extract_from_columns(df)
        
        # যদি না পাওয়া যায়, তাহলে এলোমেলো ডাটা থেকে খোঁজা
        if clean_df is None or clean_df.empty:
            clean_df = extract_from_messy_data(df)
        
        if clean_df is None or clean_df.empty:
            await update.message.reply_text("❌ Could not extract data! Make sure data is in proper format.")
            return

        # ডুপ্লিকেট চেক (সমস্ত কলাম চেক করা হবে)
        duplicate_mask = clean_df.duplicated(keep='first')
        unique_df = clean_df[~duplicate_mask]
        duplicate_df = clean_df[duplicate_mask]

        # ফাইল তৈরি
        unique_df.to_excel(unique_file, index=False)
        
        if not duplicate_df.empty:
            duplicate_df.to_excel(duplicate_file, index=False)
            auto_adjust_column_width(unique_file)
            auto_adjust_column_width(duplicate_file)
        else:
            auto_adjust_column_width(unique_file)

        # রিপোর্ট
        report = f"✅ *Results:*\n"
        report += f"📊 Original: {len(clean_df)} records\n"
        report += f"✅ Unique: {len(unique_df)} records\n"
        report += f"🗑️ Duplicates removed: {len(duplicate_df)} records\n\n"
        report += f"📌 Same passwords grouped together\n"
        report += f"📌 Columns: col1 | col2 | col3 | col4 | col5"

        # ফাইল পাঠানো
        with open(unique_file, 'rb') as f:
            await update.message.reply_document(
                document=f,
                caption=report,
                parse_mode="Markdown"
            )

        if not duplicate_df.empty:
            with open(duplicate_file, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    caption=f"🗑️ *Duplicate Records* (Total: {len(duplicate_df)})",
                    parse_mode="Markdown"
                )

    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)}\n\n⚠️ Make sure file format is correct")

    finally:
        for f in [input_file, unique_file, duplicate_file]:
            if os.path.exists(f):
                os.remove(f)

# ================= 🔄 [ মেইন ] =================
def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    print("=" * 50)
    print("✅ DUPLICATE REMOVER BOT STARTED")
    print("📌 Works with 1 or more columns")
    print("📌 If 1 column: splits by spaces")
    print("📌 Extracts up to 5 columns")
    print("📌 Groups same passwords together")
    print("💡 Send any Excel/CSV file")
    print("=" * 50)
    
    app.run_polling()

if __name__ == '__main__':
    main()

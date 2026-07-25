import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

# ================= 🔧 [ কনফিগারেশন ] =================
BOT_TOKEN = os.environ.get("BOT_TOKEN")

def auto_adjust_column_width(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(file_path)
    except:
        pass

def extract_data(df):
    all_data = []
    num_cols = df.shape[1]
    
    if num_cols == 1:
        for idx, row in df.iterrows():
            if pd.notna(row[0]):
                text = str(row[0]).strip()
                if text:
                    parts = text.split()
                    row_data = parts[:5]
                    while len(row_data) < 5:
                        row_data.append("")
                    all_data.append(row_data)
    else:
        max_cols = min(5, num_cols)
        for idx, row in df.iterrows():
            row_data = []
            for col_idx in range(max_cols):
                if pd.notna(row[col_idx]):
                    text = str(row[col_idx]).strip()
                    if text and ' ' in text:
                        parts = text.split()
                        row_data.extend(parts[:5])
                    else:
                        row_data.append(text)
                else:
                    row_data.append("")
            
            if len(row_data) > 5:
                row_data = row_data[:5]
            while len(row_data) < 5:
                row_data.append("")
            
            if row_data[0] and row_data[1]:
                all_data.append(row_data)
    
    if all_data:
        columns = ["col1", "col2", "col3", "col4", "col5"]
        df_clean = pd.DataFrame(all_data, columns=columns)
        df_clean = df_clean[(df_clean["col1"].astype(str).str.strip() != "") & 
                            (df_clean["col2"].astype(str).str.strip() != "")]
        df_clean = df_clean.sort_values(by=["col2", "col1"])
        return df_clean
    return None

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "✅ *Duplicate Remover Bot Active!*\n\n"
        "📂 Send me a `.csv` or `.xlsx` file.\n\n"
        "🔧 *What this bot does:*\n"
        "• Works with 1 or more columns\n"
        "• If 1 column: splits data by spaces\n"
        "• Extracts up to 5 columns\n"
        "• Removes duplicate entries\n"
        "• Groups same passwords together\n\n"
        "💡 Powered by MAX FUTURE",
        parse_mode="Markdown"
    )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file_name = update.message.document.file_name
    input_file = f"input_{file_name}"
    unique_file = "Unique_Data.xlsx"
    duplicate_file = "Duplicate_Data.xlsx"

    try:
        file = await context.bot.get_file(update.message.document.file_id)
        await file.download_to_drive(input_file)
        await update.message.reply_text("⏳ Processing file...")

        if file_name.lower().endswith('.csv'):
            try:
                df = pd.read_csv(input_file, encoding='utf-8', header=None)
            except:
                df = pd.read_csv(input_file, encoding='latin1', header=None)
        else:
            df = pd.read_excel(input_file, header=None)

        clean_df = extract_data(df)
        
        if clean_df is None or clean_df.empty:
            await update.message.reply_text("❌ Could not extract data! Make sure data is in proper format.")
            return

        duplicate_mask = clean_df.duplicated(keep='first')
        unique_df = clean_df[~duplicate_mask]
        duplicate_df = clean_df[duplicate_mask]

        unique_df.to_excel(unique_file, index=False)
        
        if not duplicate_df.empty:
            duplicate_df.to_excel(duplicate_file, index=False)
            auto_adjust_column_width(unique_file)
            auto_adjust_column_width(duplicate_file)
        else:
            auto_adjust_column_width(unique_file)

        report = f"✅ *Results:*\n"
        report += f"📊 Original: {len(clean_df)} records\n"
        report += f"✅ Unique: {len(unique_df)} records\n"
        report += f"🗑️ Duplicates removed: {len(duplicate_df)} records"

        with open(unique_file, 'rb') as f:
            await update.message.reply_document(
                document=f,
                caption=report,
                parse_mode="Markdown"
            )

        if not duplicate_df.empty:
            with open(duplicate_file, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    caption=f"🗑️ *Duplicate Records* (Total: {len(duplicate_df)})",
                    parse_mode="Markdown"
                )

    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)}")

    finally:
        for f in [input_file, unique_file, duplicate_file]:
            if os.path.exists(f):
                os.remove(f)

def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    print("=" * 50)
    print("✅ DUPLICATE REMOVER BOT STARTED")
    print("💡 Send any Excel/CSV file")
    print("=" * 50)
    
    app.run_polling()

if __name__ == '__main__':
    main()
